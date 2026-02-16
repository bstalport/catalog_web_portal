# -*- coding: utf-8 -*-

from odoo import http, _
from odoo.http import request
from odoo.exceptions import UserError
import csv
import io
import base64
import logging
from datetime import datetime

_logger = logging.getLogger(__name__)

# Characters that trigger formula interpretation in Excel/LibreOffice
_CSV_FORMULA_CHARS = ('=', '+', '-', '@', '\t', '\r')


def _sanitize_csv_value(value):
    """Prevent CSV formula injection (CWE-1236).

    If a cell value starts with a formula trigger character, prefix it
    with a single-quote so spreadsheet applications treat it as text.
    """
    if isinstance(value, str) and value and value[0] in _CSV_FORMULA_CHARS:
        return "'" + value
    return value


# Standard export column headers (Odoo import format)
EXPORT_HEADERS = [
    'id',                    # External ID
    'name',                  # Product Name
    'default_code',          # Internal Reference
    'barcode',               # Barcode
    'list_price',            # Sales Price
    'standard_price',        # Cost (empty — private)
    'categ_id/id',           # Category (external ID)
    'type',                  # Product Type
    'sale_ok',               # Can be Sold
    'purchase_ok',           # Can be Purchased
    'weight',                # Weight
    'volume',                # Volume
    'description_sale',      # Description
]

SUPPLIER_HEADERS = [
    'seller_ids/partner_id/id',   # Supplier partner external ID
    'seller_ids/product_code',    # Supplier's product code
    'seller_ids/product_name',    # Supplier's product name
    'seller_ids/price',           # Purchase price from supplier
    'seller_ids/min_qty',         # Minimum quantity
]


class CatalogExport(http.Controller):
    """
    Controller pour gérer les exports de catalogue.
    Supporte CSV, Excel, et import direct Odoo.
    """

    # ============ CSV EXPORT ============
    
    @http.route(['/catalog/export/csv'], 
                type='http', auth='user', methods=['POST'])
    def export_csv(self, product_ids=None, **kwargs):
        """
        Génère un fichier CSV compatible avec l'import Odoo standard.
        
        Args:
            product_ids: IDs de produits séparés par virgules (ou depuis session)
        
        Returns:
            CSV file download
        """
        try:
            # Récupérer le client
            partner = request.env.user.partner_id
            catalog_client = request.env['catalog.client'].sudo().search([
                ('partner_id', '=', partner.id),
                ('is_active', '=', True)
            ], limit=1)
            
            if not catalog_client:
                raise UserError(_('You do not have catalog access.'))
            
            # Récupérer les produits à exporter
            if product_ids:
                # IDs fournis en paramètre
                product_ids_list = [int(pid) for pid in product_ids.split(',')]
            else:
                # Sinon, depuis la session (sélection)
                product_ids_list = request.session.get('catalog_selection', [])
            
            if not product_ids_list:
                raise UserError(_('No products selected for export.'))
            
            # Vérifier la limite d'export
            config = request.env['catalog.config'].sudo().get_config()
            if config.max_products_per_export > 0:
                if len(product_ids_list) > config.max_products_per_export:
                    raise UserError(_(
                        'Maximum %s products per export. Please reduce your selection.'
                    ) % config.max_products_per_export)
            
            # Vérifier rate limiting
            if config.export_rate_limit > 0:
                recent_exports = request.env['catalog.access.log'].sudo().search_count([
                    ('client_id', '=', catalog_client.id),
                    ('action', '=', 'export_csv'),
                    ('create_date', '>=', datetime.now().replace(minute=0, second=0, microsecond=0))
                ])
                if recent_exports >= config.export_rate_limit:
                    raise UserError(_(
                        'Export limit reached (%s exports per hour). Please try again later.'
                    ) % config.export_rate_limit)
            
            # Récupérer les produits
            Product = request.env['product.template'].sudo()
            products = Product.browse(product_ids_list)
            
            # Vérifier l'accès via domain SQL
            access_domain = catalog_client._get_accessible_domain()
            products = Product.search([('id', 'in', product_ids_list)] + access_domain)
            
            if not products:
                raise UserError(_('No accessible products to export.'))
            
            # Générer le CSV
            output = io.StringIO()
            writer = csv.writer(output, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)

            # Check if supplier info should be included
            include_supplier_info = config.include_supplier_info_in_exports
            supplier_external_id = config.supplier_external_id or 'catalog_supplier'

            headers = list(EXPORT_HEADERS)
            if include_supplier_info:
                headers.extend(SUPPLIER_HEADERS)
            headers.append('image_1920')

            writer.writerow(headers)

            # Prix selon pricelist du client
            pricelist = catalog_client.pricelist_id

            # Ajouter les produits
            for product in products:
                # Calculer le prix
                if pricelist:
                    price = pricelist._get_product_price(product, 1.0)
                else:
                    price = product.list_price

                # Image (optionnel - peut être lourd)
                include_images = kwargs.get('include_images', False)
                if include_images and product.image_1920:
                    # Convertir en base64
                    image_b64 = base64.b64encode(product.image_1920).decode('utf-8')
                else:
                    image_b64 = ''

                row = [
                    f'__import__.supplier_{catalog_client.id}_product_{product.id}',  # External ID unique
                    _sanitize_csv_value(product.name),
                    _sanitize_csv_value(product.default_code or ''),
                    _sanitize_csv_value(product.barcode or ''),
                    price,
                    '',  # Coût vide (info privée fournisseur)
                    f'__import__.{product.categ_id.name.lower().replace(" ", "_")}' if product.categ_id else '',
                    product.type or 'consu',
                    'True',  # Vendable
                    'True',  # Achetable (enabled for supplier products)
                    product.weight or 0,
                    product.volume or 0,
                    _sanitize_csv_value(product.description_sale or ''),
                ]

                # Add supplier info for invoice recognition
                if include_supplier_info:
                    row.extend([
                        f'__import__.{supplier_external_id}',  # Supplier partner external ID
                        _sanitize_csv_value(product.default_code or ''),  # Supplier's product code
                        _sanitize_csv_value(product.name),                # Supplier's product name
                        price,                                 # Purchase price (same as catalog price)
                        1.0,                                   # Minimum quantity
                    ])

                # Add image
                row.append(image_b64)

                writer.writerow(row)
            
            # Logger l'export
            request.env['catalog.access.log'].sudo().log_action(
                action='export_csv',
                client_id=catalog_client.id,
                user_id=request.env.user.id,
                product_ids=products.ids,
                ip_address=request.httprequest.remote_addr,
                export_format='csv',
                success=True,
            )
            
            # Préparer la réponse
            output.seek(0)
            csv_data = output.getvalue()
            
            # Générer nom de fichier
            filename = f'catalog_export_{catalog_client.partner_id.name}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
            filename = filename.replace(' ', '_').replace('/', '_')
            
            return request.make_response(
                csv_data,
                headers=[
                    ('Content-Type', 'text/csv; charset=utf-8'),
                    ('Content-Disposition', f'attachment; filename={filename}'),
                ]
            )
        
        except Exception as e:
            # Logger l'erreur
            try:
                request.env['catalog.access.log'].sudo().log_action(
                    action='export_csv',
                    client_id=catalog_client.id if 'catalog_client' in locals() else None,
                    user_id=request.env.user.id,
                    ip_address=request.httprequest.remote_addr,
                    success=False,
                    error_message=str(e),
                )
            except:
                pass
            
            raise UserError(str(e))
    
    # ============ EXCEL EXPORT ============

    @http.route(['/catalog/export/excel'],
                type='http', auth='user', methods=['POST'])
    def export_excel(self, product_ids=None, **kwargs):
        """
        Génère un fichier Excel (.xlsx) avec mise en forme.
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
        import io as _io

        try:
            # Récupérer le client
            partner = request.env.user.partner_id
            catalog_client = request.env['catalog.client'].sudo().search([
                ('partner_id', '=', partner.id),
                ('is_active', '=', True)
            ], limit=1)

            if not catalog_client:
                raise UserError(_('You do not have catalog access.'))

            # Vérifier que l'export Excel est activé
            config = request.env['catalog.config'].sudo().get_config()
            if not config.allow_excel_export:
                raise UserError(_('Excel export is not enabled.'))

            # Récupérer les produits à exporter
            if product_ids:
                product_ids_list = [int(pid) for pid in product_ids.split(',')]
            else:
                product_ids_list = request.session.get('catalog_selection', [])

            if not product_ids_list:
                raise UserError(_('No products selected for export.'))

            # Vérifier la limite d'export
            if config.max_products_per_export > 0:
                if len(product_ids_list) > config.max_products_per_export:
                    raise UserError(_(
                        'Maximum %s products per export. Please reduce your selection.'
                    ) % config.max_products_per_export)

            # Vérifier rate limiting
            if config.export_rate_limit > 0:
                recent_exports = request.env['catalog.access.log'].sudo().search_count([
                    ('client_id', '=', catalog_client.id),
                    ('action', 'in', ['export_csv', 'export_excel']),
                    ('create_date', '>=', datetime.now().replace(minute=0, second=0, microsecond=0))
                ])
                if recent_exports >= config.export_rate_limit:
                    raise UserError(_(
                        'Export limit reached (%s exports per hour). Please try again later.'
                    ) % config.export_rate_limit)

            # Récupérer les produits
            Product = request.env['product.template'].sudo()
            products = Product.browse(product_ids_list)

            # Vérifier l'accès via domain SQL
            access_domain = catalog_client._get_accessible_domain()
            products = Product.search([('id', 'in', product_ids_list)] + access_domain)

            if not products:
                raise UserError(_('No accessible products to export.'))

            # Prix selon pricelist du client
            pricelist = catalog_client.pricelist_id
            include_images = kwargs.get('include_images', False)

            # Check if supplier info should be included
            include_supplier_info = config.include_supplier_info_in_exports
            supplier_external_id = config.supplier_external_id or 'catalog_supplier'

            # ---- Créer le workbook ----
            wb = Workbook()
            ws = wb.active
            ws.title = 'Product Catalog'

            # Styles
            header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='2E5984', end_color='2E5984', fill_type='solid')
            header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

            thin_border = Border(
                left=Side(style='thin', color='D0D0D0'),
                right=Side(style='thin', color='D0D0D0'),
                bottom=Side(style='thin', color='D0D0D0'),
                top=Side(style='thin', color='D0D0D0'),
            )

            alt_row_fill = PatternFill(start_color='F2F7FC', end_color='F2F7FC', fill_type='solid')
            data_align = Alignment(vertical='center', wrap_text=True)
            price_format = '#,##0.00'

            headers = list(EXPORT_HEADERS)
            col_widths = [32, 30, 18, 16, 14, 14, 26, 14, 12, 14, 10, 10, 36]

            if include_supplier_info:
                headers.extend(SUPPLIER_HEADERS)
                col_widths.extend([28, 18, 28, 14, 12])

            if include_images:
                headers.append('image_1920')
                col_widths.append(16)

            # ---- En-tête ----
            ws.row_dimensions[1].height = 22
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border
                ws.column_dimensions[get_column_letter(col_idx)].width = col_widths[col_idx - 1]

            # ---- Données ----
            for row_idx, product in enumerate(products, start=2):
                # Prix
                if pricelist:
                    price = pricelist._get_product_price(product, 1.0)
                else:
                    price = product.list_price

                row_data = [
                    f'__import__.supplier_{catalog_client.id}_product_{product.id}',
                    _sanitize_csv_value(product.name),
                    _sanitize_csv_value(product.default_code or ''),
                    _sanitize_csv_value(product.barcode or ''),
                    price,
                    '',  # Coût vide
                    f'__import__.{product.categ_id.name.lower().replace(" ", "_")}' if product.categ_id else '',
                    product.type or 'consu',
                    True,
                    True,  # Achetable (enabled for supplier products)
                    product.weight or 0,
                    product.volume or 0,
                    _sanitize_csv_value(product.description_sale or ''),
                ]

                # Add supplier info for invoice recognition
                if include_supplier_info:
                    row_data.extend([
                        f'__import__.{supplier_external_id}',  # Supplier partner external ID
                        _sanitize_csv_value(product.default_code or ''),  # Supplier's product code
                        _sanitize_csv_value(product.name),                # Supplier's product name
                        price,                                 # Purchase price
                        1.0,                                   # Minimum quantity
                    ])

                # Add image placeholder (will be filled below if needed)
                if include_images:
                    row_data.append('')

                is_alt = (row_idx % 2 == 0)

                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = thin_border
                    cell.alignment = data_align
                    if is_alt:
                        cell.fill = alt_row_fill

                    # Formatage numérique
                    if col_idx == 5:  # list_price
                        cell.number_format = price_format
                    elif col_idx == 6:  # standard_price
                        cell.number_format = price_format
                    elif col_idx in (11, 12):  # Weight, Volume
                        cell.number_format = '0.000'
                    # Supplier price column
                    if include_supplier_info and col_idx == 17:  # seller_ids/price
                        cell.number_format = price_format

                # Image en base64 dans la dernière colonne (texte)
                if include_images:
                    img_col = len(row_data)
                    img_cell = ws.cell(row=row_idx, column=img_col)
                    if product.image_1920:
                        img_cell.value = base64.b64encode(product.image_1920).decode('utf-8')
                    img_cell.border = thin_border
                    img_cell.alignment = data_align
                    if is_alt:
                        img_cell.fill = alt_row_fill

            # ---- Auto-filter sur les en-têtes ----
            last_col_letter = get_column_letter(len(headers))
            last_row = len(products) + 1
            ws.auto_filter.ref = f'A1:{last_col_letter}{last_row}'

            # ---- Freeze panes sous l'en-tête ----
            ws.freeze_panes = 'A2'

            # ---- Onglet "Info" avec métadonnées ----
            ws_info = wb.create_sheet('Export Info')
            info_data = [
                ['Export Details', ''],
                ['', ''],
                ['Generated by', 'Catalog Web Portal'],
                ['Client', catalog_client.partner_id.name],
                ['Date', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
                ['Products exported', len(products)],
                ['Pricelist', pricelist.name if pricelist else 'Default'],
            ]

            # Add supplier info section if enabled
            if include_supplier_info:
                company = request.env.company
                info_data.extend([
                    ['', ''],
                    ['SUPPLIER INFO FOR INVOICE RECOGNITION', ''],
                    ['', ''],
                    ['Supplier External ID', f'__import__.{supplier_external_id}'],
                    ['Supplier Name', company.name],
                    ['', ''],
                    ['IMPORTANT: Before importing this file, you must:', ''],
                    ['1. Create a supplier partner in your Odoo', ''],
                    ['2. Set its External ID to:', f'__import__.{supplier_external_id}'],
                    ['', ''],
                    ['How to set External ID:', ''],
                    ['- Enable Developer Mode (Settings > Activate Developer Mode)', ''],
                    ['- Open the partner form', ''],
                    ['- Use Debug menu > View Metadata > External ID', ''],
                    ['- Or import a partner CSV with the "id" column', ''],
                    ['', ''],
                    ['After import, your invoices from this supplier', ''],
                    ['will automatically match products by reference.', ''],
                ])

            title_font = Font(name='Calibri', size=13, bold=True, color='2E5984')
            label_font = Font(name='Calibri', size=11, bold=True)
            warning_font = Font(name='Calibri', size=11, bold=True, color='D9534F')

            for r_idx, (label, value) in enumerate(info_data, start=1):
                ws_info.cell(row=r_idx, column=1, value=label)
                ws_info.cell(row=r_idx, column=2, value=value)
                if r_idx == 1:
                    ws_info.cell(row=r_idx, column=1).font = title_font
                elif label == 'SUPPLIER INFO FOR INVOICE RECOGNITION':
                    ws_info.cell(row=r_idx, column=1).font = title_font
                elif label == 'IMPORTANT: Before importing this file, you must:':
                    ws_info.cell(row=r_idx, column=1).font = warning_font
                elif r_idx > 2 and label:
                    ws_info.cell(row=r_idx, column=1).font = label_font

            ws_info.column_dimensions['A'].width = 50
            ws_info.column_dimensions['B'].width = 36

            # ---- Logger l'export ----
            request.env['catalog.access.log'].sudo().log_action(
                action='export_excel',
                client_id=catalog_client.id,
                user_id=request.env.user.id,
                product_ids=products.ids,
                ip_address=request.httprequest.remote_addr,
                export_format='excel',
                success=True,
            )

            # ---- Sérialiser ----
            buffer = _io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            filename = f'catalog_export_{catalog_client.partner_id.name}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            filename = filename.replace(' ', '_').replace('/', '_')

            return request.make_response(
                buffer.getvalue(),
                headers=[
                    ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                    ('Content-Disposition', f'attachment; filename={filename}'),
                ]
            )

        except Exception as e:
            try:
                request.env['catalog.access.log'].sudo().log_action(
                    action='export_excel',
                    client_id=catalog_client.id if 'catalog_client' in locals() else None,
                    user_id=request.env.user.id,
                    ip_address=request.httprequest.remote_addr,
                    success=False,
                    error_message=str(e),
                )
            except:
                pass

            raise UserError(str(e))
    
    # ============ DIRECT ODOO IMPORT (Future) ============
    
    @http.route(['/catalog/export/direct'], 
                type='json', auth='user', methods=['POST'])
    def export_direct_odoo(self, product_ids, odoo_url, db, username, password, **kwargs):
        """
        Import direct dans Odoo client via XML-RPC.
        
        Args:
            product_ids: Liste d'IDs de produits
            odoo_url: URL de l'instance Odoo client
            db: Nom de la base de données
            username: Login Odoo
            password: Mot de passe
        
        Returns:
            dict: {success: bool, imported: int, errors: list}
        """
        import xmlrpc.client
        
        try:
            # Récupérer le client
            partner = request.env.user.partner_id
            catalog_client = request.env['catalog.client'].sudo().search([
                ('partner_id', '=', partner.id),
                ('is_active', '=', True)
            ], limit=1)
            
            if not catalog_client:
                return {'success': False, 'error': 'No catalog access'}
            
            # Vérifier que la fonctionnalité est activée
            config = request.env['catalog.config'].sudo().get_config()
            if not config.allow_direct_odoo_import:
                return {'success': False, 'error': 'Direct import is disabled'}
            
            # Connexion XML-RPC au Odoo client
            common = xmlrpc.client.ServerProxy(f'{odoo_url}/xmlrpc/2/common')
            uid = common.authenticate(db, username, password, {})
            
            if not uid:
                return {'success': False, 'error': 'Authentication failed'}
            
            models = xmlrpc.client.ServerProxy(f'{odoo_url}/xmlrpc/2/object')
            
            # Récupérer les produits
            Product = request.env['product.template'].sudo()
            products = Product.browse(product_ids)
            
            # Vérifier accès via domain SQL
            access_domain = catalog_client._get_accessible_domain()
            products = Product.search([('id', 'in', product_ids)] + access_domain)
            
            if not products:
                return {'success': False, 'error': 'No accessible products'}
            
            # Prix selon pricelist
            pricelist = catalog_client.pricelist_id
            
            # Importer les produits
            imported = 0
            errors = []
            
            for product in products:
                try:
                    # Calculer prix
                    if pricelist:
                        price = pricelist._get_product_price(product, 1.0)
                    else:
                        price = product.list_price
                    
                    # External ID pour éviter doublons
                    external_id = f'supplier_{catalog_client.id}_product_{product.id}'
                    
                    # Vérifier si existe déjà
                    existing = models.execute_kw(
                        db, uid, password,
                        'product.template', 'search',
                        [[('default_code', '=', product.default_code)]]
                    ) if product.default_code else []
                    
                    vals = {
                        'name': product.name,
                        'default_code': product.default_code,
                        'list_price': price,
                        'type': product.type,
                        'barcode': product.barcode,
                        'weight': product.weight,
                        'volume': product.volume,
                        'description_sale': product.description_sale,
                    }
                    
                    # Image (optionnel)
                    if kwargs.get('include_images') and product.image_1920:
                        vals['image_1920'] = base64.b64encode(product.image_1920).decode('utf-8')
                    
                    if existing:
                        # Update
                        models.execute_kw(
                            db, uid, password,
                            'product.template', 'write',
                            [existing, vals]
                        )
                    else:
                        # Create
                        models.execute_kw(
                            db, uid, password,
                            'product.template', 'create',
                            [vals]
                        )
                    
                    imported += 1
                
                except Exception as e:
                    errors.append(f'{product.name}: {str(e)}')
            
            # Logger
            request.env['catalog.access.log'].sudo().log_action(
                action='direct_import',
                client_id=catalog_client.id,
                user_id=request.env.user.id,
                product_ids=products.ids,
                ip_address=request.httprequest.remote_addr,
                success=True,
                action_details=f'Imported {imported} products to {odoo_url}',
            )
            
            return {
                'success': True,
                'imported': imported,
                'total': len(products),
                'errors': errors
            }
        
        except Exception as e:
            return {'success': False, 'error': str(e)}
